from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains as action
from datetime import date
from pathlib import Path
import inspect
import pytest
import random
import openpyxl
from constants import globalConstants as C



class Test_sauce:

    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.get(C.base_URL)
        self.driver.maximize_window()
        self.folderpath = str(date.today())
        Path(self.folderpath).mkdir(exist_ok=True)
        

    def Getdata():
        excel_file = openpyxl.load_workbook("Data\Homework_datas.xlsx")
        selected_shet = excel_file["Sayfa1"]
        totalrows = selected_shet.max_row
        Datas = []
        for i in range(2,totalrows +1):
            kullaniciad = selected_shet.cell(i,1).value
            sifre = selected_shet.cell(i,2).value 
            tupleData = (kullaniciad,sifre)
            Datas.append(tupleData)
            return Datas

    def teardown (self):
        self.driver.quit()

    def wait_for_element(self,locator):
        WebDriverWait(self.driver,5).until(EC.visibility_of_element_located(locator))


    def test_empty_username_password(self):
        self.wait_for_element((By.ID, C.lgnbtn_id))
        lgnbtn = self.driver.find_element(By.ID, C.lgnbtn_id).click()
        self.wait_for_element((By.XPATH, C.username_required_msg_xpath))
        required_message = self.driver.find_element(By.XPATH, C.username_required_msg_xpath)
        self.driver.save_screenshot(f"{self.folderpath}\{inspect.currentframe().f_code.co_name}.png")
        assert required_message.text == C.expexted_username_required_msg

    def test_empty_password(self):
        self.wait_for_element((By.ID, C.username_id))
        username_input = self.driver.find_element(By.ID, C.username_id)
        action(self.driver).move_to_element(username_input).click().send_keys("Merve").perform()
        self.wait_for_element((By.ID, C.lgnbtn_id))
        lgnbtn = self.driver.find_element(By.ID, C.lgnbtn_id).click()
        required_message = self.driver.find_element(By.XPATH, C.password_required_msg_xpath)
        self.driver.save_screenshot(f"{self.folderpath}\{inspect.currentframe().f_code.co_name}.png")
        assert required_message.text == C.expexted_password_required_msg

    def test_locked_out(self):
        self.wait_for_element((By.ID, C.username_id ))
        username_input = self.driver.find_element(By.ID, C.username_id)
        action(self.driver).move_to_element(username_input).click().send_keys(C.locked_username).perform()
        self.wait_for_element((By.ID, C.password_id))
        password_input = self.driver.find_element(By.ID, C.password_id)
        action(self.driver).move_to_element(password_input).click().send_keys(C.password).perform()
        self.wait_for_element((By.ID, C.lgnbtn_id))
        lgnbtn = self.driver.find_element(By.ID, C.lgnbtn_id ).click()
        locked_out_message = self.driver.find_element(By.XPATH, C.locked_user_msg_xpath)
        self.driver.save_screenshot(f"{self.folderpath}\{inspect.currentframe().f_code.co_name}.png")
        assert locked_out_message.text == C.expexted_locked_user_msg
        

    def test_icon_x(self):
        try:
            icon_username = self.driver.find_element(By.CSS_SELECTOR, C.icon_x_selector1) # 1 vve 2 silinince ikisini de buluyo mu kontrol et.
            icon_password = self.driver.find_element(By.CSS_SELECTOR, C.icon_x_selector)
        except:
            True
        self.wait_for_element((By.ID, C.lgnbtn_id))
        lgnbtn = self.driver.find_element(By.ID, C.lgnbtn_id).click()
        icon_username = self.driver.find_element(By.CSS_SELECTOR, C.icon_x_selector)
        icon_password = self.driver.find_element(By.CSS_SELECTOR, C.icon_x_selector1)
        if icon_username.is_displayed() and icon_password.is_displayed():
            True
        else:
            False #Bosken logine bastik. ikonlar çikmadi sikinti var.
        self.driver.find_element(By.CSS_SELECTOR, C.lgn_btn_x_button ).click()
        try:
            icon_username.is_displayed()
            icon_password.is_displayed()
            False
        except:
            True #("ikonlar tekrar gitti her şey yolunda")
        self.driver.save_screenshot(f"{self.folderpath}\{inspect.currentframe().f_code.co_name}.png")
        



    def test_Page(self):
            self.wait_for_element((By.ID, C.standard_username))
            username = self.driver.find_element(By.ID, C.username_id)
            action(self.driver).move_to_element(username).click().send_keys(C.standard_username).perform()
            self.wait_for_element((By.ID, C.password_id))
            password =self.driver.find_element(By.ID, C.password_id)
            action(self.driver).move_to_element(password).click().send_keys(C.password).perform()
            self.wait_for_element((By.ID, C.lgnbtn_id))
            self.driver.find_element(By.ID, C.lgnbtn_id).click()
            WebDriverWait(self.driver,5).until(EC.new_window_is_opened)
            Current_url = self.driver.current_url
            self.driver.save_screenshot(f"{self.folderpath}\{inspect.currentframe().f_code.co_name}.png")
            assert C.inventory_URL == Current_url



            

    def test_Does_Inventory_have_six_items(self):
        self.wait_for_element((By.ID, C.username_id))
        username = self.driver.find_element(By.ID, C.username_id)
        action(self.driver).move_to_element(username).click().send_keys(C.standard_username).perform()
        self.wait_for_element((By.ID, C.password_id))
        password= self.driver.find_element(By.ID, C.password_id)
        action(self.driver).move_to_element(password).click().send_keys(C.password).perform()
        self.wait_for_element((By.ID, C.lgnbtn_id))
        self.driver.find_element(By.ID, C.lgnbtn_id).click()
        WebDriverWait(self.driver,10).until(EC.new_window_is_opened)
        products = self.driver.find_elements(By.CLASS_NAME, C.inventory_items_class_name)
        self.driver.save_screenshot(f"{self.folderpath}\{inspect.currentframe().f_code.co_name}.png")
        assert len(products) == 6


    def test_add_to_cart(self):
        self.wait_for_element((By.ID, C.username_id))
        username_input = self.driver.find_element(By.ID, C.username_id)
        action(self.driver).move_to_element(username_input).click().send_keys(C.problem_username).perform()
        self.wait_for_element((By.ID, C.password_id))
        password_input = self.driver.find_element(By.ID, C.password_id)
        action(self.driver).move_to_element(password_input).click().send_keys(C.password).perform()
        self.wait_for_element((By.ID, C.lgnbtn_id))
        lgnbtn = self.driver.find_element(By.ID, C.lgnbtn_id).click()
        self.wait_for_element((By. ID, C.inventory_container_id))
        add_to_cart = self.driver.find_element(By.ID, "add-to-cart-sauce-labs-backpack")
        add_to_cart.click()
        sepet = self.driver.find_element(By.XPATH, '//*[@id="shopping_cart_container"]/a')
        sepet.click()
        self.driver.save_screenshot(f"{self.folderpath}\{inspect.currentframe().f_code.co_name}.png")
        assert self.driver.find_element(By.XPATH, '//*[@id="cart_contents_container"]/div/div[1]/div[3]')

    @pytest.mark.xfail()
    def test_image_ctrl(self):
        self.wait_for_element((By.ID, C.username_id))
        username_input = self.driver.find_element(By.ID, C.username_id)
        action(self.driver).move_to_element(username_input).click().send_keys(C.problem_username).perform()
        self.wait_for_element((By.ID, C.password_id))
        password_input = self.driver.find_element(By.ID, C.password_id)
        action(self.driver).move_to_element(password_input).click().send_keys(C.password).perform()
        self.wait_for_element((By.ID, C.lgnbtn_id))
        lgnbtn = self.driver.find_element(By.ID, C.lgnbtn_id).click()
        self.wait_for_element((By. ID, C.inventory_container_id))
        entry_images = self.driver.find_elements(By.XPATH, C.inventory_items_imgs_xpath)
        entry_images_url_list=[]
        for i in entry_images:
                img_url = i.get_attribute("src")
                entry_images_url_list.append(img_url)
        num_of_img = len(entry_images)
        second_images_url_list = []
        for i in range(num_of_img):
            entry_images = self.driver.find_elements(By.XPATH, C.inventory_items_imgs_xpath)
            WebDriverWait(self.driver,2)
            entry_images[i].click()
            WebDriverWait(self.driver,2)
            second_img = self.driver.find_element(By.CLASS_NAME, C.inventory_details_img_class_name)
            second_img_url = second_img.get_attribute("src")
            second_images_url_list.append(second_img_url)
            self.driver.find_element(By.ID, C.back_products_btn_id).click()
        self.driver.save_screenshot(f"{self.folderpath}\{inspect.currentframe().f_code.co_name}.png")
        for i in range(num_of_img):
            assert entry_images_url_list[i] == second_images_url_list[i]

    @pytest.mark.xfail()
    def test_different_user(self):
        self.wait_for_element((By.ID, C.username_id))
        username_input = self.driver.find_element(By.ID, C.username_id)
        action(self.driver).move_to_element(username_input).click().send_keys(C.problem_username).perform()
        self.wait_for_element((By.ID, C.password_id))
        password_input = self.driver.find_element(By.ID, C.password_id)
        action(self.driver).move_to_element(password_input).click().send_keys(C.password).perform()
        self.wait_for_element((By.ID, C.lgnbtn_id))
        lgnbtn = self.driver.find_element(By.ID, C.lgnbtn_id).click()
        self.wait_for_element((By. ID, C.inventory_container_id))
        add_to_cart = self.driver.find_element(By.ID, "add-to-cart-sauce-labs-backpack").click()
        burger_button = self.driver.find_element(By.ID, C.burger_btn_id).click()
        logout_button = self.driver.find_element(By.ID, C.logout_btn_id).click()
        self.wait_for_element((By.ID, C.username_id))
        username_input = self.driver.find_element(By.ID, C.username_id)
        action(self.driver).move_to_element(username_input).click().send_keys(C.standard_username).perform()
        self.wait_for_element((By.ID, C.password_id))
        password_input = self.driver.find_element(By.ID, C.password_id)
        action(self.driver).move_to_element(password_input).click().send_keys(C.password).perform()
        self.wait_for_element((By.ID, C.lgnbtn_id))
        lgnbtn = self.driver.find_element(By.ID, C.lgnbtn_id).click()
        self.wait_for_element((By. ID, C.inventory_container_id))
        basket_content = self.driver.find_element(By.XPATH, C.basket_content_container_xpath )
        basket_content.click()
        self.driver.save_screenshot(f"{self.folderpath}\{inspect.currentframe().f_code.co_name}.png")
        assert not self.driver.find_element(By.XPATH, '//*[@id="cart_contents_container"]/div/div[1]/div[3]')



    @pytest.mark.parametrize("user,secret",[("standard_user","secret_sauce"),
        ("performance_glitch_user","secret_sauce"),
        pytest.param("locked_out_user","secret_sauce", marks=pytest.mark.xfail),
        pytest.param("problem_user","secret_sauce", marks=pytest.mark.xfail)
    ])
    def test_product_num_in_basket(self,user,secret):
        self.wait_for_element((By.ID, C.username_id))
        username_input = self.driver.find_element(By.ID, C.username_id)
        action(self.driver).move_to_element(username_input).click().send_keys(user).perform()
        self.wait_for_element((By.ID, "password"))
        password_input = self.driver.find_element(By.ID, C.password_id)
        action(self.driver).move_to_element(password_input).click().send_keys(secret).perform()
        self.wait_for_element((By.ID, C.lgnbtn_id))
        lgnbtn = self.driver.find_element(By.ID, C.lgnbtn_id).click()
        self.wait_for_element((By. ID, C.inventory_container_id))
        self.wait_for_element((By. XPATH, C.add_cart_btns_xpath))
        add_cart_buttons = self.driver.find_elements(By. XPATH, C.add_cart_btns_xpath)
        random_num= random.randint(1,6)
        random_num= int(random_num)
        index = 0
        for i in range(random_num): 
            add_cart_buttons = self.driver.find_elements(By. XPATH, C.add_cart_btns_xpath)
            add_cart_buttons[index].click()
            WebDriverWait(self.driver,2)
            index+=1
        basket= self.driver.find_element(By.CLASS_NAME, C.shop_cart_badge_class_name)
        self.driver.save_screenshot(f"{self.folderpath}\{inspect.currentframe().f_code.co_name}.png")
        assert basket.text == str(random_num)

    def test_total_price_ctrl(self):
        self.wait_for_element((By.ID, C.username_id))
        username_input = self.driver.find_element(By.ID, C.username_id)
        action(self.driver).move_to_element(username_input).click().send_keys(C.standard_username).perform()
        self.wait_for_element((By.ID, C.password_id))
        password_input = self.driver.find_element(By.ID, C.password_id)
        action(self.driver).move_to_element(password_input).click().send_keys(C.password).perform()
        self.wait_for_element((By.ID, C.lgnbtn_id))
        lgnbtn = self.driver.find_element(By.ID, C.lgnbtn_id).click()
        self.wait_for_element((By. ID, C.inventory_container_id))
        self.wait_for_element((By. XPATH, C.add_cart_btns_xpath))
        add_cart_buttons = self.driver.find_elements(By. XPATH, C.add_cart_btns_xpath)
        prices = self.driver.find_elements(By.XPATH, C.item_prices_xpath)
        price_list =[]
        for i in prices:
            price = i.text.replace("$", "")
            price_list.append(price)
        add_cart_buttons_list = []
        for button in add_cart_buttons:
            add_cart_buttons_list.append(button)
        button_price_list = dict(zip(add_cart_buttons_list,price_list))
        random_num= random.randint(1,6)
        random_num= int(random_num)
        index = 0
        total_value = 0
        for i in range(random_num): 
            add_cart_buttons = self.driver.find_elements(By. XPATH, C.add_cart_btns_xpath)
            add_cart_buttons[index].click()
            WebDriverWait(self.driver,2)
            total_value += float(price_list[index])
            index+=1
        self.driver.find_element(By.CLASS_NAME, C.shop_cart_badge_class_name).click() #alışveriş sepeti görseli
        self.wait_for_element((By.ID, C.checkout_btn_id))
        self.driver.find_element(By.ID, C.checkout_btn_id).click() #checkout button
        name = self.driver.find_element(By.ID, C.checkout_name_id)
        last_name= self.driver.find_element(By.ID, C.checkout_lastname_id)
        zip_code= self.driver.find_element(By.ID, C.checkout_zip_id)
        action(self.driver).move_to_element(name).click().send_keys(C.checkout_name).perform()
        action(self.driver).move_to_element(last_name).click().send_keys(C.checkout_lastname).perform()
        action(self.driver).move_to_element(zip_code).click().send_keys(C.checkout_zip).perform()
        self.driver.find_element(By.ID, "continue").click()
        self.wait_for_element((By.XPATH, C.total_price_box_xpath))
        total_movey_window= self.driver.find_element(By.XPATH,C.total_price_box_xpath)
        tax= self.driver.find_element(By.XPATH, C.tax_box_xpath)
        tax_float = float(tax.text.replace("Tax: $",""))
        items_total_on_last_page = float(total_movey_window.text.replace("Total: $",""))
        tax_plus_value = round(total_value+tax_float,2)
        self.driver.save_screenshot(f"{self.folderpath}\{inspect.currentframe().f_code.co_name}.png")
        assert tax_plus_value == items_total_on_last_page



    @pytest.mark.parametrize("kullaniciadi,sifre",Getdata())
    def test_none_user(self, kullaniciadi, sifre):
        self.wait_for_element((By.ID, C.username_id ))
        username_input = self.driver.find_element(By.ID, C.username_id)
        action(self.driver).move_to_element(username_input).click().send_keys(kullaniciadi).perform()
        self.wait_for_element((By.ID, C.password_id))
        password_input = self.driver.find_element(By.ID, C.password_id)
        action(self.driver).move_to_element(password_input).click().send_keys(sifre).perform()
        self.wait_for_element((By.ID, C.lgnbtn_id))
        self.driver.find_element(By.ID, C.lgnbtn_id ).click()
        none_user_message = self.driver.find_element(By.XPATH, C.none_user_mgs_xpath)
        self.driver.save_screenshot(f"{self.folderpath}\{inspect.currentframe().f_code.co_name}.png")
        assert none_user_message.text == C.expected_none_user_msg

    def test_twitter_btn_ctrl(self):
        self.wait_for_element((By.ID, C.username_id ))
        username_input = self.driver.find_element(By.ID, C.username_id)
        action(self.driver).move_to_element(username_input).click().send_keys(C.standard_username).perform()
        self.wait_for_element((By.ID, C.password_id))
        password_input = self.driver.find_element(By.ID, C.password_id)
        action(self.driver).move_to_element(password_input).click().send_keys(C.password).perform()
        self.wait_for_element((By.ID, C.lgnbtn_id))
        self.driver.find_element(By.ID, C.lgnbtn_id ).click()
        self.wait_for_element((By.XPATH, C.twitter_btn_xpath))
        self.driver.find_element((By.XPATH, C.twitter_btn_xpath)).click()
        Current_url = self.driver.current_url
        self.driver.save_screenshot(f"{self.folderpath}\{inspect.currentframe().f_code.co_name}.png")
        assert Current_url == C.expected_twitter_url

    def test_fb_btn_ctrl(self):
        self.wait_for_element((By.ID, C.username_id ))
        username_input = self.driver.find_element(By.ID, C.username_id)
        action(self.driver).move_to_element(username_input).click().send_keys(C.standard_username).perform()
        self.wait_for_element((By.ID, C.password_id))
        password_input = self.driver.find_element(By.ID, C.password_id)
        action(self.driver).move_to_element(password_input).click().send_keys(C.password).perform()
        self.wait_for_element((By.ID, C.lgnbtn_id))
        self.driver.find_element(By.ID, C.lgnbtn_id ).click()
        self.wait_for_element((By.XPATH, C.fb_btn_xpath))
        self.driver.find_element((By.XPATH, C.fb_btn_xpath)).click()
        Current_url = self.driver.current_url
        self.driver.save_screenshot(f"{self.folderpath}\{inspect.currentframe().f_code.co_name}.png")
        assert Current_url == C.expected_fb_url

    def test_linkedin_btn_ctrl(self):
        self.wait_for_element((By.ID, C.username_id ))
        username_input = self.driver.find_element(By.ID, C.username_id)
        action(self.driver).move_to_element(username_input).click().send_keys(C.standard_username).perform()
        self.wait_for_element((By.ID, C.password_id))
        password_input = self.driver.find_element(By.ID, C.password_id)
        action(self.driver).move_to_element(password_input).click().send_keys(C.password).perform()
        self.wait_for_element((By.ID, C.lgnbtn_id))
        self.driver.find_element(By.ID, C.lgnbtn_id ).click()
        self.wait_for_element((By.XPATH, C.linkedin_btn_xpath))
        self.driver.find_element((By.XPATH, C.linkedin_btn_xpath)).click()
        Current_url = self.driver.current_url
        self.driver.save_screenshot(f"{self.folderpath}\{inspect.currentframe().f_code.co_name}.png")
        assert Current_url == C.expected_linkedin_url

    def test_item_names(self):
        self.wait_for_element((By.ID, C.username_id))
        username = self.driver.find_element(By.ID, C.username_id)
        action(self.driver).move_to_element(username).click().send_keys(C.standard_username).perform()
        self.wait_for_element((By.ID, C.password_id))
        password= self.driver.find_element(By.ID, C.password_id)
        action(self.driver).move_to_element(password).click().send_keys(C.password).perform()
        self.wait_for_element((By.ID, C.lgnbtn_id))
        self.driver.find_element(By.ID, C.lgnbtn_id).click()
        WebDriverWait(self.driver,10).until(EC.new_window_is_opened)
        item_names = self.driver.find_elements(By.CLASS_NAME, C.items_name_class_name)
        for item in item_names:
            assert item.text.startswith("Sauce Labs")
        self.driver.save_screenshot(f"{self.folderpath}\{inspect.currentframe().f_code.co_name}.png")



    





    

            


        






































            